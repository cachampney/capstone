"""
Module containing budget class
"""


import json

import transaction
from goal import Goal
from transaction import Transaction
import openpyxl


class Budget:
    """
    Object representing a budget. Holds transactions and other budget related data.
    """

    def __init__(self):
        self.expense_transactions = []
        self.income_transactions = []
        self.categories = ["Utilities", "Gas", "Entertainment", "Rent/Housing",
                           "Groceries", "Other"]
        self.expense_goals = {}  # name of goal is key and Goal object is value
        self.total_expenses = 0
        self.total_income = 0
        self.balance = 0

    def add_transaction(self, transaction: Transaction):
        """
        Method to add transaction to budget

        :param transaction: transaction to add to budget
        :type transaction: Transaction
        :return: bool - True if transaction added successful, False if not
        """
        # validate transaction
        if not transaction or not isinstance(transaction, Transaction):
            raise ValueError("Transaction must not be empty and must be of type Transaction")
        if transaction.transaction_type.lower() == 'income':
            self.income_transactions.append(transaction)
            self.total_income += transaction.amount
        if transaction.transaction_type.lower() == 'expense':
            self.expense_transactions.append(transaction)
            self.total_expenses += transaction.amount
        self.balance = self.total_income - self.total_expenses  # update balance
        if transaction.expense_goal != "N/A":
            self.link_transaction_to_expense_goal(transaction, transaction.expense_goal)
        return True

    def delete_transaction(self, transaction: Transaction):
        """
        Method to delete transaction from budget

        :param transaction: transaction to delete
        :type transaction: Transaction
        :return: bool - True if transaction removed, False if not (likely transaction was not in list)
        """
        if not transaction or not isinstance(transaction, Transaction):
            raise ValueError("Transaction must not be empty and must be of type Transaction")
        if transaction.transaction_type.lower() == 'income':
            try:
                self.income_transactions.remove(transaction)
            except ValueError:
                return False
            else:
                self.total_income -= transaction.amount
        if transaction.transaction_type.lower() == 'expense':
            try:
                self.expense_transactions.remove(transaction)
            except ValueError:
                return False
            else:
                self.total_expenses -= transaction.amount
        self.balance = self.total_income - self.total_expenses  # update balance
        if transaction.expense_goal != "N/A":
            self.unlink_transaction_from_expense_goal(transaction, transaction.expense_goal)
        return True

    def new_transaction_update_goal_amount(self, transaction :Transaction):
        """
        Method to update amounts in the goal transaction

        :param transaction: transaction to update the goal amount
        """
        goal = self.get_expense_goal(transaction.expense_goal)

        if transaction.transaction_type == "Expense":
            goal.calc_currentAmount(transaction.amount)
            return goal
        else:
            goal.remove_currentAmount(transaction.amount)
            return goal

    def update_transaction_expense_goal_cell(self, transaction :Transaction):
        """
        Method to update a cell and goal when the ender user updates the expense goal column

        :param transaction: transaction to update a cell in the transaction table
        """
        goal = self.get_expense_goal(transaction.expense_goal)

        if transaction.transaction_type == "Expense":
            goal.remove_currentAmount(transaction.amount)
            return goal
        else:
            goal.calc_currentAmount(transaction.amount)
            return goal
    def get_transactions(self, **kwargs):
        """
        Method to get list of transactions based on specified attribute values
        Example: budget.get_transactions(transaction_type="income", "amount"=58.25) would return
         list of transactions that are income transactions and have amount of 58.25.

        :param kwargs: keyword arguments (key=value) to get transactions based on
        :return: [Transaction] - list of Transactions, empty if no matching transactions found
        """
        trans_to_return = []  # empty list to hold transactions to return
        valid_keys = ['transaction_type', 'date', 'amount', 'vendor', 'category', 'note']
        # validate passed in keyword arguments
        for key in kwargs.keys():
            if key not in valid_keys:
                raise ValueError(f"Invalid key ({key}) provided. Valid keys are {valid_keys}.")
        for trans in self.expense_transactions:  # search through expense transactions
            for attribute, value in kwargs.items():
                if trans.get(attribute) == value:
                    trans_to_return.append(trans)
        for trans in self.income_transactions:  # search through income transactions
            for attribute, value in kwargs.items():
                if trans.get(attribute) == value:
                    trans_to_return.append(trans)
        return trans_to_return

    def add_category(self, category: str):
        """
        method to add a category to the budget

        :param category: name of category to add
        :type category: str
        :return: None
        """
        if not category or not isinstance(category, str):
            raise ValueError("Category must be non-empty string")
        self.categories.append(category)

    def remove_category(self, category: str):
        """
        method to remove a category from the budget

        :param category: name of category to add
        :type category: str
        :return: None
        """
        if not category or not isinstance(category, str):
            raise ValueError("Category must be non-empty string")
        try:
            self.categories.remove(category)
        except ValueError:
            print(f"Could not delete Category {category}. It was not found in the budget")

    def add_expense_goal(self, goal: Goal):
        """
        method to add an expanse goal to the budget
        storing goal in self.expense_goals dict
        key is lowercase version of goal name for efficiency

        :param goal: name of category to add
        :type goal: Goal
        :return: None
        """
        if not goal or not isinstance(goal, Goal):
            raise ValueError("Goal must be of type Goal")
        if goal.name.lower() in self.expense_goals:
            raise ValueError(f"Goal with name {goal.name} already exists in budget")
        self.expense_goals[goal.name.lower()] = goal

    def get_expense_goal(self, goal_name: str):
        """
        Method to get goal from Budget for specified name

        :param goal_name: name of goal to retrieve from budget
        :type goal_name: str
        :return: Goal - desired expense goal
        """
        try:
            return self.expense_goals[goal_name.lower()]
        except KeyError:
            raise ValueError(f"goal with name {goal_name} not found in budget")

    def delete_expense_goal(self, goal_name: str):
        """
        Method to delete goal with specified name from Budget

        :param goal_name: name of goal to delete from budget
        :type goal_name: str
        :return: None
        """
        try:
            del self.expense_goals[goal_name.lower()]
        except KeyError:
            raise ValueError(f"Could not delete goal {goal_name}. It was not found in the budget")

    def save_budget(self, filename):
        """
        Method to save all budget data to specified file

        :param filename: filename, including path, for save file
        :type filename: str
        :return: None
        """
        budget_dict = {'total_expenses': self.total_expenses, 'total_income': self.total_income,
                       'balance': self.balance, 'expense_transactions': [], 'income_transactions': [],
                       'goals': []}
        for trans in self.expense_transactions:
            budget_dict['expense_transactions'].append(trans.to_dict())
        for trans in self.income_transactions:
            budget_dict['income_transactions'].append(trans.to_dict())
        budget_dict['categories'] = self.categories
        for goal in self.expense_goals.values():
            budget_dict['goals'].append(goal.to_dict())
        write_json_to_file(budget_dict, filename)

    def load_budget(self, filename):
        """
        Method to load all budget data from specified file

        :param filename: filename, including path, for save file
        :type filename: str
        :return: None
        """
        budget_dict = read_json_file(filename)
        self.categories = budget_dict['categories']
        for trans_dict in budget_dict['expense_transactions']:
            t = Transaction()
            t.update_from_dict(trans_dict)
            self.add_transaction(t)
        for trans_dict in budget_dict['income_transactions']:
            t = Transaction()
            t.update_from_dict(trans_dict)
            self.add_transaction(t)
        for goal_dict in budget_dict['goals']:
            g = Goal()
            g.update_from_dict(goal_dict)
            self.add_expense_goal(g)

    def delete_budget(self, filename):
        """
        Function to delete a saved budget

        :param filename: full filepath of budget save file
        :type filename: str
        :return: None
        """
        if os.path.exists(filename):
            os.remove(filename)

    def link_transaction_to_expense_goal(self, transaction, expense_goal_name):
        """
        Method to link transaction to a goal

        :param transaction: transaction to apply to goal
        :type transaction: Transaction
        :param expense_goal_name: Name of expense goal to link transaction to
        :type expense_goal_name: str
        :return: None
        """
        try:
            self.expense_goals['expense_goal_name'].apply_transaction(transaction)
        except KeyError:
            print('goal does not exist')

    def unlink_transaction_from_expense_goal(self, transaction, expense_goal_name):
        """
        Method to remove transaction from expense goal

        :param transaction: transaction to apply to goal
        :type transaction: Transaction
        :param expense_goal_name: Name of expense goal to link transaction to
        :type expense_goal_name: str
        :return: None
        """
        try:
            self.expense_goals['expense_goal_name'].remove_transaction(transaction)
        except KeyError:
            print('goal does not exist in budget')

    def export_transactions(self, filename):
        """
        Method to write income and expense transactions to Excel file

        :param filename: filename, including path, for save file
        :type filename: str
        :return: None
        """
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        transaction_objs = self.expense_transactions + self.income_transactions
        transaction_dicts = []
        for trans in transaction_objs:
            transaction_dicts.append(trans.to_dict())
        headers = list(transaction_dicts[0].keys())
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)
        for row_num, row_data in enumerate(transaction_dicts, 2):
            for col_num, key in enumerate(headers, 1):
                sheet.cell(row=row_num, column=col_num, value=row_data[key])
        workbook.save(filename)


def write_json_to_file(json_data, filename):
    """
    Function to write json encode-able data to specified file

    :param json_data: data to write to file
    :type json_data: str|dict|list|int
    :param filename: filename, including path, for save file
    :type filename: str
    :return: None
    """
    with open(filename, 'w') as json_file:
        json.dump(json_data, json_file)


def read_json_file(filename):
    """
    Function to open json file and read in data as dictionary

    :param filename: filename, including path, for save file
    :type filename: str
    :return: dict - data from file
    :rtype: dict
    """
    with open(filename, 'r') as json_file:
        data = json.load(json_file)
    return data
